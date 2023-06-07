from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook

insu = 90000
url  = f'https://www.nps.or.kr/jsppage/app/etc/simpleExpect.jsp?yyyy=2022&status=cal&max=471600&min=29700&insu={insu}&inquiry=%BF%B9%BB%F3%BF%AC%B1%DD%BE%D7+%C1%B6%C8%B8%C7%CF%B1%E2'
req = requests.get(url)
req.encoding = "utf-8"
html = BeautifulSoup(req.content,"html.parser")

years10 = html.find('dt', class_='th1_1 th1').find_next_sibling().find('span').text
years15 = html.find('dt', class_='th1_2 th1').find_next_sibling().find('span').text
years20 = html.find('dt', class_='th1_3 th1').find_next_sibling().find('span').text
years25 = html.find('dt', class_='th1_4 th1').find_next_sibling().find('span').text
years30 = html.find('dt', class_='th1_5 th1').find_next_sibling().find('span').text
years35 = html.find('dt', class_='th1_6 th1').find_next_sibling().find('span').text
years40 = html.find('dt', class_='th1_7 th1').find_next_sibling().find('span').text


a = [years10, years15, years20, years25, years30, years35, years40]
print(f"노령연금은{a}")

#이 코드는 html 변수에 할당된 BeautifulSoup 객체 안에서 <dt> 태그 중 class 속성이 "th1_1 th1"인 태그를 찾고, 그 후에 그 태그의 다음 요소로 위치한 <dd> 태그를 찾고, 그 안에 있는 <span> 태그를 찾아 그 안에 있는 값을 추출합니다.
#참고로 이 코드는 <dt> 태그가 하나만 존재할 경우에만 정상적으로 작동합니다. 여러 개의 <dt> 태그가 존재할 경우에는 첫 번째 <dt> 태그의 값만 추출됩니다. 여러 개의 <dt> 태그 중 원하는 값이 있는 태그를 찾기 위해서는 추가적인 절차가 필요할 수 있습니다.

wb = Workbook()
ws = wb.active

ws.append([f"노령연금 입금금액 = {insu}"])
ws.append(["10년","15년","20년","25년","30년","35년","40년"])
ws['A3'] = years10 
ws['B3'] = years15 
ws['C3'] = years20 
ws['D3'] = years25 
ws['E3'] = years30 
ws['F3'] = years35 
ws['G3'] = years40 

wb.save('yungum.xlsx')
wb.close()
