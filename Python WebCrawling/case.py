from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook

insu = 330000
url  = f'https://www.nps.or.kr/jsppage/app/etc/simpleExpect.jsp?yyyy=2022&status=cal&max=471600&min=29700&insu={insu}&inquiry=%BF%B9%BB%F3%BF%AC%B1%DD%BE%D7+%C1%B6%C8%B8%C7%CF%B1%E2'
req = requests.get(url)
req.encoding = "utf-8"
html = BeautifulSoup(req.content,"html.parser")

data_1 = html.select("span.data")
years10 = data_1[0].string
years15 = data_1[1].string
years20 = data_1[2].string
years25 = data_1[3].string
years30 = data_1[4].string
years35 = data_1[5].string
years40 = data_1[6].string


a = [years10, years15, years20, years25, years30, years35, years40]
print(f"노령연금은{a}")

#select로 웹크롤링할때 원하는페이지에서 웹페이지 검사로 span.data에서 원하는 수치가 때문에 
#span.data에서 크롤링하였다. .string은 불필요한 태그가 나오는것을 방지하기 위해서 넣었다. 
#insu는 최소 소득기준이 330000이기 때문이다.

wb = Workbook()
ws = wb.active

ws.append([f"노령연금 입금금액 = {insu}"])
ws.append(["10년","15년","20년","25년","30년","35년","40년"])
ws['A3'] = years10 
ws['B3'] = years15 
ws['C3'] = years20 
ws['D3'] = years25 
ws['E3'] = years30 
ws['F3'] = years35 
ws['G3'] = years40 

wb.save('yungum.xlsx')
wb.close()
